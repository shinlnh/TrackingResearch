import argparse
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import torch

from pytracking.analysis.plot_results import check_and_load_precomputed_results
from pytracking.evaluation import Tracker, get_dataset


SUCCESS_THRESHOLD_INDEX = 10   # 0.50 on 0:0.05:1
OP75_THRESHOLD_INDEX = 15      # 0.75 on 0:0.05:1
PRECISION_THRESHOLD_INDEX = 20
NORM_PRECISION_THRESHOLD_INDEX = 20


def read_time_stats(tracker_results_dir: Path, seq_name: str) -> tuple[float, int, float]:
    time_path = tracker_results_dir / f"{seq_name}_time.txt"
    if not time_path.exists():
        return float("nan"), 0, float("nan")

    times = np.loadtxt(time_path, delimiter="\t")
    times = np.atleast_1d(times).astype(np.float64)
    total_time = float(np.sum(times))
    frames = int(times.shape[0])
    fps = float(frames / total_time) if total_time > 0 else float("nan")
    return fps, frames, total_time


def round_or_nan(value: float, digits: int = 4):
    return value if np.isnan(value) else round(float(value), digits)


def build_rows(dataset, eval_data, tracker_results_dir: Path):
    valid = torch.tensor(eval_data["valid_sequence"], dtype=torch.bool)
    avg_overlap_all = torch.tensor(eval_data["avg_overlap_all"], dtype=torch.float64)
    success_overlap = torch.tensor(eval_data["ave_success_rate_plot_overlap"], dtype=torch.float64)
    precision_curve = torch.tensor(eval_data["ave_success_rate_plot_center"], dtype=torch.float64)

    rows = []
    missing_sequences = []

    for seq_id, seq in enumerate(dataset):
        if not valid[seq_id]:
            missing_sequences.append(seq.name)
            continue

        fps, frames, total_time = read_time_stats(tracker_results_dir, seq.name)
        rows.append(
            {
                "sequence": seq.name,
                "AUC": float(avg_overlap_all[seq_id, 0] * 100.0),
                "Precision": float(precision_curve[seq_id, 0, PRECISION_THRESHOLD_INDEX] * 100.0),
                "Success": float(success_overlap[seq_id, 0, SUCCESS_THRESHOLD_INDEX] * 100.0),
                "FPS": fps,
                "frames": frames,
                "total_time_sec": total_time,
            }
        )

    return rows, missing_sequences


def compute_summary(rows, eval_data, tracker_name: str, parameter_name: str, run_id: int, missing_sequences):
    success_overlap = torch.tensor(eval_data["ave_success_rate_plot_overlap"], dtype=torch.float64)
    precision_curve = torch.tensor(eval_data["ave_success_rate_plot_center"], dtype=torch.float64)
    norm_precision_curve = torch.tensor(eval_data["ave_success_rate_plot_center_norm"], dtype=torch.float64)
    valid = torch.tensor(eval_data["valid_sequence"], dtype=torch.bool)

    success_overlap = success_overlap[valid].mean(0) * 100.0
    precision_curve = precision_curve[valid].mean(0) * 100.0
    norm_precision_curve = norm_precision_curve[valid].mean(0) * 100.0

    auc_mean = float(torch.tensor([row["AUC"] for row in rows], dtype=torch.float64).mean())
    precision_mean = float(precision_curve[0, PRECISION_THRESHOLD_INDEX])
    norm_precision_mean = float(norm_precision_curve[0, NORM_PRECISION_THRESHOLD_INDEX])
    success_mean = float(success_overlap[0, SUCCESS_THRESHOLD_INDEX])
    op75 = float(success_overlap[0, OP75_THRESHOLD_INDEX])

    fps_values = np.array([row["FPS"] for row in rows], dtype=np.float64)
    total_frames = int(sum(row["frames"] for row in rows))
    total_time = float(sum(row["total_time_sec"] for row in rows))
    fps_avg_seq = float(np.nanmean(fps_values))
    fps_median_seq = float(np.nanmedian(fps_values))
    fps_weighted_by_frames = float(total_frames / total_time) if total_time > 0 else float("nan")

    summary = {
        "tracker": f"{tracker_name}_{parameter_name}_{run_id}",
        "dataset": "OTB100",
        "valid_sequences": f"{len(rows)}/{len(rows) + len(missing_sequences)}",
        "missing_sequences": ", ".join(missing_sequences),
        "AUC_mean": auc_mean,
        "Precision_mean": precision_mean,
        "Norm_Precision_mean": norm_precision_mean,
        "Success_mean": success_mean,
        "OP75_mean": op75,
        "FPS_avg_seq": fps_avg_seq,
        "FPS_median_seq": fps_median_seq,
        "FPS_global": fps_avg_seq,
        "FPS_weighted_by_frames": fps_weighted_by_frames,
        "total_frames": total_frames,
        "total_time_sec": total_time,
    }
    return summary


def write_metrics_workbook(path: Path, rows, summary, include_overall_row: bool):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "per_sequence"
    ws.append(("sequence", "AUC", "Precision", "Success", "FPS", "frames", "total_time_sec"))

    for row in rows:
        ws.append(
            (
                row["sequence"],
                round(row["AUC"], 4),
                round(row["Precision"], 4),
                round(row["Success"], 4),
                round_or_nan(row["FPS"], 4),
                row["frames"],
                round_or_nan(row["total_time_sec"], 6),
            )
        )

    if include_overall_row:
        ws.append(
            (
                "OVERALL",
                round(summary["AUC_mean"], 4),
                round(summary["Precision_mean"], 4),
                round(summary["Success_mean"], 4),
                round_or_nan(summary["FPS_global"], 4),
                summary["total_frames"],
                round_or_nan(summary["total_time_sec"], 6),
            )
        )

    ws2 = wb.create_sheet("summary")
    ws2.append(("metric", "value"))
    ws2.append(("tracker", summary["tracker"]))
    ws2.append(("dataset", summary["dataset"]))
    ws2.append(("valid_sequences", summary["valid_sequences"]))
    ws2.append(("missing_sequences", summary["missing_sequences"]))
    ws2.append(("AUC_mean", round(summary["AUC_mean"], 4)))
    ws2.append(("Precision_mean", round(summary["Precision_mean"], 4)))
    ws2.append(("Success_mean", round(summary["Success_mean"], 4)))
    ws2.append(("FPS_avg_seq", round_or_nan(summary["FPS_avg_seq"], 4)))
    ws2.append(("FPS_median_seq", round_or_nan(summary["FPS_median_seq"], 4)))
    ws2.append(("FPS_global", round_or_nan(summary["FPS_global"], 4)))
    ws2.append(("FPS_weighted_by_frames", round_or_nan(summary["FPS_weighted_by_frames"], 4)))

    wb.save(path)


def save_scatter_plot(path: Path, rows):
    auc = np.array([row["AUC"] for row in rows], dtype=np.float64)
    fps = np.array([row["FPS"] for row in rows], dtype=np.float64)

    fig, ax = plt.subplots(figsize=(8, 6))
    sc = ax.scatter(fps, auc, c=auc, cmap="viridis", alpha=0.8, edgecolors="black", linewidths=0.3)
    ax.set_title("AUC vs FPS")
    ax.set_xlabel("FPS")
    ax.set_ylabel("AUC")
    ax.grid(True, alpha=0.3)
    fig.colorbar(sc, ax=ax, label="AUC")
    fig.tight_layout()
    fig.savefig(path, dpi=200)
    plt.close(fig)


def save_distribution_plot(path: Path, rows):
    metrics = [
        ("AUC", "AUC"),
        ("Precision", "Precision"),
        ("Success", "Success"),
        ("FPS", "FPS"),
    ]

    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    for ax, (key, title) in zip(axes.flatten(), metrics):
        values = np.array([row[key] for row in rows], dtype=np.float64)
        ax.hist(values[np.isfinite(values)], bins=15, color="#2a6f97", alpha=0.85, edgecolor="white")
        ax.set_title(title)
        ax.grid(True, alpha=0.2)
    fig.tight_layout()
    fig.savefig(path, dpi=200)
    plt.close(fig)


def save_sorted_plot(path: Path, rows):
    metrics = [
        ("AUC", "AUC"),
        ("Precision", "Precision"),
        ("Success", "Success"),
        ("FPS", "FPS"),
    ]

    fig, axes = plt.subplots(2, 2, figsize=(13, 8))
    for ax, (key, title) in zip(axes.flatten(), metrics):
        values = np.array(sorted((row[key] for row in rows), reverse=True), dtype=np.float64)
        ax.plot(np.arange(1, len(values) + 1), values, linewidth=1.8, color="#bc4749")
        ax.set_title(f"{title} Sorted Desc")
        ax.set_xlabel("Sequence rank")
        ax.set_ylabel(title)
        ax.grid(True, alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=200)
    plt.close(fig)


def save_sequence_order_plot(path: Path, rows):
    metrics = [
        ("AUC", "AUC"),
        ("Precision", "Precision"),
        ("Success", "Success"),
        ("FPS", "FPS"),
    ]

    x = np.arange(1, len(rows) + 1)
    fig, axes = plt.subplots(4, 1, figsize=(16, 10), sharex=True)
    for ax, (key, title) in zip(axes, metrics):
        values = np.array([row[key] for row in rows], dtype=np.float64)
        ax.plot(x, values, linewidth=1.2, color="#386641")
        ax.set_ylabel(title)
        ax.grid(True, alpha=0.25)
    axes[0].set_title("Metrics by Sequence Order")
    axes[-1].set_xlabel("Sequence index")
    fig.tight_layout()
    fig.savefig(path, dpi=200)
    plt.close(fig)


def update_benchmark_workbook(path: Path, profile: str, run_id: int, summary, notes: str):
    header = (
        "timestamp",
        "profile",
        "run_id",
        "dataset",
        "valid_sequences",
        "AUC",
        "Precision",
        "Norm_Precision",
        "OP50",
        "OP75",
        "Global_FPS",
        "Avg_Seq_FPS",
        "Median_Seq_FPS",
        "notes",
    )

    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()

    ws = wb["OTB100"] if "OTB100" in wb.sheetnames else wb.active
    ws.title = "OTB100"
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.append(header)
    elif tuple(cell.value for cell in ws[1][: len(header)]) != header:
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.delete_rows(1, 1)
        ws.insert_rows(1)
        for col, value in enumerate(header, start=1):
            ws.cell(row=1, column=col, value=value)

    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 2).value == profile and ws.cell(row_idx, 3).value == run_id:
            target_row = row_idx
            break
    if target_row is None:
        target_row = ws.max_row + 1

    values = (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        profile,
        run_id,
        summary["dataset"],
        summary["valid_sequences"],
        round(summary["AUC_mean"], 2),
        round(summary["Precision_mean"], 2),
        round(summary["Norm_Precision_mean"], 2),
        round(summary["Success_mean"], 2),
        round(summary["OP75_mean"], 2),
        round(summary["FPS_global"], 4),
        round(summary["FPS_avg_seq"], 4),
        round(summary["FPS_median_seq"], 4),
        notes,
    )
    for col, value in enumerate(values, start=1):
        ws.cell(target_row, col, value)

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Build OTB run summary workbooks and plots.")
    parser.add_argument("--tracker-name", required=True)
    parser.add_argument("--parameter-name", required=True)
    parser.add_argument("--run-id", required=True, type=int)
    parser.add_argument("--dataset-name", default="otb")
    parser.add_argument("--report-dir", required=True)
    parser.add_argument("--benchmark-xlsx", default="")
    args = parser.parse_args()

    report_dir = Path(args.report_dir).resolve()
    report_dir.mkdir(parents=True, exist_ok=True)

    dataset = get_dataset(args.dataset_name)
    tracker = Tracker(args.tracker_name, args.parameter_name, args.run_id)
    report_name = f"{args.dataset_name}_{args.tracker_name}_{args.parameter_name}_{args.run_id:03d}"
    eval_data = check_and_load_precomputed_results(
        [tracker], dataset, report_name, force_evaluation=True, skip_missing_seq=False, verbose=False
    )

    rows, missing_sequences = build_rows(dataset, eval_data, Path(tracker.results_dir))
    summary = compute_summary(rows, eval_data, args.tracker_name, args.parameter_name, args.run_id, missing_sequences)

    report_prefix = report_dir.name
    write_metrics_workbook(report_dir / f"{report_prefix}_sequence_metrics.xlsx", rows, summary, include_overall_row=False)
    write_metrics_workbook(
        report_dir / f"{report_prefix}_sequence_metrics_with_overall.xlsx",
        rows,
        summary,
        include_overall_row=True,
    )

    save_scatter_plot(report_dir / "auc_vs_fps_scatter.png", rows)
    save_distribution_plot(report_dir / "metrics_distribution.png", rows)
    save_sorted_plot(report_dir / "metrics_sorted_desc.png", rows)
    save_sequence_order_plot(report_dir / "metrics_by_sequence_order.png", rows)

    if args.benchmark_xlsx:
        notes = ""
        if missing_sequences:
            notes = f"Missing sequence: {', '.join(missing_sequences)}"
        update_benchmark_workbook(Path(args.benchmark_xlsx), args.parameter_name, args.run_id, summary, notes)

    print(f"Report created in {report_dir}")
    print(
        f"AUC={summary['AUC_mean']:.4f}, Precision={summary['Precision_mean']:.4f}, "
        f"Success={summary['Success_mean']:.4f}, FPS_global={summary['FPS_global']:.4f}"
    )


if __name__ == "__main__":
    main()
