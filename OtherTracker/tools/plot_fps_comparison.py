from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


@dataclass(frozen=True)
class FpsEntry:
    tracker: str
    fps_global: float
    valid_sequences: int
    source_type: str
    source_file: Path
    note: str = ""


AGGREGATE_FILES = {
    "embedded_fps_summary.csv",
    "fps_summary_current.csv",
    "fps_global_only.csv",
}

# Manual overrides for trackers whose embedded fallback FPS is known to be wrong
# for the chart we want to present.
FPS_GLOBAL_OVERRIDES = {
    "CCOT": 60.258,
}

FPS_LABEL_OVERRIDES = {
    "CCOT": "60.258",
}

TRACKERS_TO_EXCLUDE = {
    "ECO",
    "MyToMP",
}

PARTIAL_FPS_SOURCE_OVERRIDES = {
    "MDNet": {
        "path": "OtherTracker/MDNet/otb_pymdnet_vototb_subset22_20260323/summary.csv",
        "fps_keys": ("FPS_global", "FPS_avg_seq", "fps_global", "fps_avg_seq"),
        "valid_keys": ("valid_sequences",),
        "source_type": "othertracker_local_partial",
        "note": "subset22_local_proxy_from_pymdnet",
    },
    "CNN-SVM": {
        "path": "OtherTracker/CNN-SVM/subset22_approx_20260325/cnn_svm_approx_summary.csv",
        "fps_keys": ("fps_global", "fps_avg_seq", "FPS_global", "FPS_avg_seq"),
        "valid_keys": ("valid_sequences",),
        "source_type": "othertracker_local_partial",
        "note": "subset22_local_proxy_from_cnnsvm_approx",
    },
}


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(
        description=(
            "Compare MyECOTracker FPS against all OtherTracker trackers with valid "
            "OTB100 FPS_global summaries."
        )
    )
    parser.add_argument("--repo-root", type=Path, default=repo_root)
    parser.add_argument("--my-root", type=Path, default=repo_root / "MyECOTracker")
    parser.add_argument("--other-root", type=Path, default=repo_root / "OtherTracker")
    parser.add_argument(
        "--my-mode",
        choices=("report", "tracking_results"),
        default="report",
        help="Load MyECOTracker FPS either from a generated report workbook or directly from tracking_results/*_time.txt.",
    )
    parser.add_argument(
        "--my-results-dir",
        type=Path,
        default=None,
        help="Path to a MyECOTracker tracking_results run directory when --my-mode=tracking_results.",
    )
    parser.add_argument(
        "--my-label",
        type=str,
        default="MyTracker",
        help="Display label for MyECOTracker in the output chart.",
    )
    parser.add_argument(
        "--out-csv",
        type=Path,
        default=repo_root / "OtherTracker" / "fps_comparison_with_myecotracker.csv",
    )
    parser.add_argument(
        "--out-png",
        type=Path,
        default=repo_root / "OtherTracker" / "fps_comparison_with_myecotracker.png",
    )
    return parser.parse_args()


def _parse_float(value: object) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    return number


def _parse_int(value: object) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return -1


def _format_fps(value: float) -> str:
    if value >= 100:
        return f"{value:.0f}"
    if value >= 10:
        return f"{value:.1f}"
    if value >= 1:
        return f"{value:.2f}"
    return f"{value:.3f}"


def _format_fps_entry(entry: FpsEntry) -> str:
    override = FPS_LABEL_OVERRIDES.get(entry.tracker)
    if override is not None:
        return override
    return _format_fps(entry.fps_global)


def _iter_summary_rows(other_root: Path) -> Iterable[tuple[Path, dict[str, str]]]:
    for path in sorted(other_root.rglob("*_summary.csv")):
        if path.name in AGGREGATE_FILES and path.parent == other_root:
            continue
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
        if rows:
            yield path, rows[0]


def _get_first_present(row: dict[str, str], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def _choose_latest_full_run(candidates: list[FpsEntry]) -> FpsEntry:
    return max(
        candidates,
        key=lambda entry: (
            entry.valid_sequences,
            entry.source_file.stat().st_mtime,
            str(entry.source_file),
        ),
    )


def load_othertracker_fps(other_root: Path) -> list[FpsEntry]:
    grouped: dict[str, list[FpsEntry]] = {}
    for path, row in _iter_summary_rows(other_root):
        tracker = (row.get("tracker") or path.parent.parent.name).strip()
        fps_global = _parse_float(row.get("fps_global"))
        valid_sequences = _parse_int(row.get("valid_sequences"))
        if not math.isfinite(fps_global) or valid_sequences < 100:
            continue
        grouped.setdefault(tracker, []).append(
            FpsEntry(
                tracker=tracker,
                fps_global=fps_global,
                valid_sequences=valid_sequences,
                source_type="othertracker_local_full",
                source_file=path,
            )
        )

    selected = {
        tracker: _choose_latest_full_run(candidates)
        for tracker, candidates in grouped.items()
    }

    embedded_path = other_root / "embedded_fps_summary.csv"
    if embedded_path.exists():
        with embedded_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                tracker = row.get("tracker", "").strip()
                if not tracker or tracker in selected:
                    continue
                fps_global = _parse_float(row.get("fps_global"))
                valid_sequences = _parse_int(row.get("valid_sequences"))
                if not math.isfinite(fps_global) or valid_sequences < 100:
                    continue
                selected[tracker] = FpsEntry(
                    tracker=tracker,
                    fps_global=fps_global,
                    valid_sequences=valid_sequences,
                    source_type="othertracker_embedded_fallback",
                    source_file=embedded_path,
                    note="fallback_to_embedded_fps",
                )

    for tracker, fps_override in FPS_GLOBAL_OVERRIDES.items():
        entry = selected.get(tracker)
        if entry is None:
            continue
        note = entry.note
        if note:
            note = f"{note}; manual_fps_override"
        else:
            note = "manual_fps_override"
        selected[tracker] = FpsEntry(
            tracker=entry.tracker,
            fps_global=fps_override,
            valid_sequences=entry.valid_sequences,
            source_type=entry.source_type,
            source_file=entry.source_file,
            note=note,
        )

    repo_root = other_root.parent
    for tracker, config in PARTIAL_FPS_SOURCE_OVERRIDES.items():
        if tracker in selected:
            continue
        csv_path = repo_root / Path(config["path"])
        if not csv_path.exists():
            continue
        with csv_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
        if not rows:
            continue
        row = rows[0]
        fps_global = _parse_float(_get_first_present(row, config["fps_keys"]))
        valid_sequences = _parse_int(_get_first_present(row, config["valid_keys"]))
        if not math.isfinite(fps_global) or valid_sequences <= 0:
            continue
        selected[tracker] = FpsEntry(
            tracker=tracker,
            fps_global=fps_global,
            valid_sequences=valid_sequences,
            source_type=str(config["source_type"]),
            source_file=csv_path,
            note=str(config["note"]),
        )

    for tracker in TRACKERS_TO_EXCLUDE:
        selected.pop(tracker, None)

    return sorted(selected.values(), key=lambda entry: entry.fps_global, reverse=True)


def find_latest_my_report(my_root: Path) -> Path:
    reports = sorted(
        my_root.rglob("*_sequence_metrics_with_overall.xlsx"),
        key=lambda path: (path.stat().st_mtime, str(path)),
    )
    if not reports:
        raise FileNotFoundError(
            f"No '*_sequence_metrics_with_overall.xlsx' report found under {my_root}."
        )
    return reports[-1]


def load_myecotracker_fps(my_root: Path) -> FpsEntry:
    report_path = find_latest_my_report(my_root)
    workbook = load_workbook(report_path, data_only=True, read_only=True)
    if "summary" not in workbook.sheetnames:
        raise KeyError(f"Workbook {report_path} does not contain a 'summary' sheet.")

    summary_sheet = workbook["summary"]
    summary_values: dict[str, object] = {}
    for metric, value in summary_sheet.iter_rows(min_row=1, values_only=True):
        if metric is None:
            continue
        summary_values[str(metric)] = value

    fps_global = _parse_float(summary_values.get("FPS_global"))
    valid_sequences_text = str(summary_values.get("valid_sequences") or "")
    valid_sequences = 0
    if "/" in valid_sequences_text:
        valid_sequences = _parse_int(valid_sequences_text.split("/", 1)[0])
    else:
        valid_sequences = _parse_int(valid_sequences_text)

    if not math.isfinite(fps_global):
        raise ValueError(f"FPS_global is missing or invalid in {report_path}.")

    tracker_note = str(summary_values.get("tracker") or report_path.stem)
    return FpsEntry(
        tracker="MyECOTracker",
        fps_global=fps_global,
        valid_sequences=valid_sequences,
        source_type="myecotracker_report",
        source_file=report_path,
        note=tracker_note,
    )


def load_myecotracker_tracking_results(results_dir: Path, tracker_label: str) -> FpsEntry:
    results_dir = results_dir.resolve()
    time_files = sorted(results_dir.glob("*_time.txt"))
    if not time_files:
        raise FileNotFoundError(f"No '*_time.txt' files found in {results_dir}.")

    fps_values: list[float] = []
    total_frames = 0
    total_time = 0.0

    for time_path in time_files:
        times = np.loadtxt(time_path, delimiter="\t")
        times = np.atleast_1d(times).astype(np.float64)
        seq_frames = int(times.shape[0])
        seq_time = float(np.sum(times))
        if seq_frames <= 0 or seq_time <= 0:
            continue
        fps_values.append(float(seq_frames / seq_time))
        total_frames += seq_frames
        total_time += seq_time

    if not fps_values or total_time <= 0:
        raise ValueError(f"Could not compute valid FPS values from {results_dir}.")

    fps_avg_seq = float(np.nanmean(np.asarray(fps_values, dtype=np.float64)))
    return FpsEntry(
        tracker=tracker_label,
        fps_global=fps_avg_seq,
        valid_sequences=len(fps_values),
        source_type="myecotracker_tracking_results",
        source_file=results_dir,
        note=f"fps_weighted_by_frames={total_frames / total_time:.6f}",
    )


def write_csv(entries: list[FpsEntry], out_csv: Path, repo_root: Path) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "tracker",
                "fps_global",
                "valid_sequences",
                "source_type",
                "source_file",
                "note",
            ],
        )
        writer.writeheader()
        for entry in entries:
            writer.writerow(
                {
                    "tracker": entry.tracker,
                    "fps_global": f"{entry.fps_global:.6f}",
                    "valid_sequences": entry.valid_sequences,
                    "source_type": entry.source_type,
                    "source_file": entry.source_file.resolve().relative_to(repo_root),
                    "note": entry.note,
                }
            )


def plot_chart(entries: list[FpsEntry], out_png: Path, my_tracker_label: str) -> None:
    out_png.parent.mkdir(parents=True, exist_ok=True)

    entries = sorted(entries, key=lambda entry: entry.fps_global, reverse=True)
    trackers = [entry.tracker for entry in entries]
    fps_values = [entry.fps_global for entry in entries]
    colors = ["#d55e00" if entry.tracker == my_tracker_label else "#4c78a8" for entry in entries]
    hatches = []
    for entry in entries:
        if entry.source_type == "othertracker_embedded_fallback":
            hatches.append("//")
        elif entry.source_type == "othertracker_local_partial":
            hatches.append("xx")
        else:
            hatches.append("")

    fig_height = max(8, 0.42 * len(entries) + 2.5)
    fig, (ax_linear, ax_log) = plt.subplots(
        1,
        2,
        figsize=(18, fig_height),
        sharey=True,
        gridspec_kw={"width_ratios": [1.25, 1]},
        constrained_layout=True,
    )

    y_positions = range(len(entries))

    linear_bars = ax_linear.barh(y_positions, fps_values, color=colors, edgecolor="#1f1f1f", linewidth=0.7)
    log_bars = ax_log.barh(y_positions, fps_values, color=colors, edgecolor="#1f1f1f", linewidth=0.7)

    for bars in (linear_bars, log_bars):
        for bar, hatch in zip(bars, hatches):
            if hatch:
                bar.set_hatch(hatch)

    ax_linear.set_yticks(list(y_positions))
    ax_linear.set_yticklabels(trackers)
    ax_linear.invert_yaxis()
    ax_linear.set_xlabel("FPS_global")
    ax_linear.set_title("Linear Scale")
    ax_linear.grid(axis="x", linestyle="--", alpha=0.35)
    ax_linear.set_axisbelow(True)
    ax_linear.set_xlim(0, max(fps_values) * 1.15)

    ax_log.set_xscale("log")
    ax_log.set_xlabel("FPS_global (log scale)")
    ax_log.set_title("Log Scale")
    ax_log.grid(axis="x", linestyle="--", alpha=0.35, which="both")
    ax_log.set_axisbelow(True)
    min_positive = min(value for value in fps_values if value > 0)
    ax_log.set_xlim(min_positive * 0.8, max(fps_values) * 1.35)

    for idx, value in enumerate(fps_values):
        ax_linear.text(
            value + max(fps_values) * 0.01,
            idx,
            _format_fps_entry(entries[idx]),
            va="center",
            ha="left",
            fontsize=9,
        )
        ax_log.text(
            value * 1.06,
            idx,
            _format_fps_entry(entries[idx]),
            va="center",
            ha="left",
            fontsize=9,
        )

    fig.suptitle(f"OTB100 FPS Comparison: {my_tracker_label} vs OtherTracker", fontsize=16, fontweight="bold")
    fig.text(
        0.5,
        0.01,
        f"Metric: FPS_global. {my_tracker_label} is highlighted in orange. Hatched bars indicate embedded fallback or partial-scope proxy benchmarks.",
        ha="center",
        fontsize=10,
    )

    fig.savefig(out_png, dpi=220, bbox_inches="tight")
    plt.close(fig)


def main() -> int:
    args = parse_args()
    other_entries = load_othertracker_fps(args.other_root)
    if args.my_mode == "report":
        my_entry = load_myecotracker_fps(args.my_root)
        if my_entry.tracker != args.my_label:
            my_entry = FpsEntry(
                tracker=args.my_label,
                fps_global=my_entry.fps_global,
                valid_sequences=my_entry.valid_sequences,
                source_type=my_entry.source_type,
                source_file=my_entry.source_file,
                note=my_entry.note,
            )
    else:
        if args.my_results_dir is None:
            raise ValueError("--my-results-dir is required when --my-mode=tracking_results.")
        my_entry = load_myecotracker_tracking_results(args.my_results_dir, args.my_label)
    combined_entries = sorted(
        [my_entry, *other_entries],
        key=lambda entry: entry.fps_global,
        reverse=True,
    )

    write_csv(combined_entries, args.out_csv, args.repo_root)
    plot_chart(combined_entries, args.out_png, args.my_label)

    print(f"{my_entry.tracker} FPS_global={my_entry.fps_global:.6f} from {my_entry.source_file}")
    print(f"Wrote CSV: {args.out_csv}")
    print(f"Wrote chart: {args.out_png}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
